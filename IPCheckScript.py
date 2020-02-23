#!./bin/python3

# -*- coding: ascii -*-

#import
import sys
import openpyxl 
#get_column_letter
#column_index_from_string
import paramiko 
import scp
import datetime
import base64

excelsheetpath = ""
excelsheet = None
sheets = {}
nodeinfo = {}
ipinfo = {}

def getWorkBookFile():
	print("IPCheck started....")
	print("Search for Excel Sheet...")
	try:  
		print("Try to use ExcelSheet given from input parameter $1: "+str(sys.argv[1]))
		curresult = str(sys.argv[1])
	except:
		print("There is no excel defined via command line, try to use standard one, ./IPCheck.xlsx.")
		curresult = "IPCheck.xlsx"

	print("Script will use: "+curresult)
	return curresult	

def getWorkbook(path):
	wb = openpyxl.load_workbook(filename = path) 
	print("Excel Sheet opened")
	return wb

def getSheets(excelsheet):
	print("Read Sheets from Workbook.")
	if len(excelsheet.get_sheet_names()) < 2:
		raise "At least two sheets must be exists: IPs, Nodes"
	
	sheets = {}
	for idx,sheet in enumerate(excelsheet.get_sheet_names()):
		sheets[sheet] = excelsheet.worksheets[idx]
	
	return sheets

def readNodes(ws):
	print("Read Nodes from Worksheet Nodes.")
	nodesinfo = {}
	loop = "true"
	idx = 2
	while loop == "true":
		nodesinfo[ws["B"+str(idx)].value] = []
		for idx2 in range(1,8):
				nodesinfo[ws["B"+str(idx)].value].append(ws[openpyxl.utils.get_column_letter(idx2)+str(idx)].value)
			
		if ws["A"+str(idx+1)].value == None:
			loop = "false"
		else:
			idx = idx + 1
	
	print("Reading Nodes finished")	
	print(nodeinfo)
	return(nodesinfo)
	
def readIPs(ws, nodecount):
	print("Read IPs from Worksheet IPs.")
	ips = {}
	loop = "true"
	idx = 2
	while loop == "true":
		ips[str(ws["B"+str(idx)].value)+':'+str(ws["C"+str(idx)].value)] = []
		for idx2 in range(1,6):
			ips[str(ws["B"+str(idx)].value)+':'+str(ws["C"+str(idx)].value)].append(ws[openpyxl.utils.get_column_letter(idx2)+str(idx)].value)

		for idx3 in range(nodecount):
			ips[str(ws["B"+str(idx)].value)+':'+str(ws["C"+str(idx)].value)].append('')

		if ws["A"+str(idx+1)].value == None:
			loop = "false"
		else:
			idx = idx + 1
	
	print("Reading IPs finished")			
	return(ips)
	
	
def writeNodes(wb, cur_sheets, cur_nodeinfo):
	print("Write NodeSheets, include Summary")
	print(cur_sheets)
	print(cur_nodeinfo)
	print(str(cur_sheets.keys()))
	for idx,node in enumerate(cur_nodeinfo):
		print(node)
		if str(node) not in str(cur_sheets.keys()):
			cur_sheets[node] = wb.create_sheet(str(node)[:30],len(cur_sheets.keys())+1)
			cur_sheets[node]["A1"] = "IP:PORT"
			cur_sheets[node]["B1"] = "Current Check Date"
			cur_sheets[node]["C1"] = "Current Check Result"
			cur_sheets[node]["D1"] = "Current Check Comment"
			cur_sheets[node]["E1"] = "Last Check Date"
			cur_sheets[node]["F1"] = "Last Check Result"
			cur_sheets[node]["G1"] = "Last Check Comment"
		cur_sheets["IPs"][openpyxl.utils.get_column_letter(idx+7)+'1'] = node 
	
	print("Write NodeSheets, include Summary done")	
	return cur_sheets

def writeNodesIPs(sheets, ips, nodes):
	#getCurrent IPs from Node
	def getNodesIPs(node):
		print("InnerFunction getNodesIPs from :"+str(node))
		nodeips = {}
		loop = "true"
		idx = 2
		while loop == "true":
			if node["A"+str(idx)].value != None:
				nodeips[str(node["A"+str(idx)].value)] = idx
				idx = idx + 1
			else:
				loop = "false"

		return nodeips
	
	#Dict, IPs von Node
	nodeips = {}
	
	#Write IPs to NodeSheets
	for node in nodes:
		#print("#############################")
		print(node)
		range = nodes[node][0]
		#print(range)
		nodeips[node] = getNodesIPs(sheets[node])
		cur_lastipcell = 2+len(nodeips[node])
		#print(cur_lastipcell)
		for ip in ips:
			print(ip)
			#print(ips[ip][0] if '|' not in ips[ip][0] else ips[ip][0].split('|')[0])
			#print(range)
			IPrange = ips[ip][0] if '|' not in ips[ip][0] else ips[ip][0].split('|')[0]
			if (ip not in str(nodeips[node].keys()) and IPrange == range):
				print("IP not found, add to excel")
				sheets[node]["A"+str(cur_lastipcell)] = ip
				nodeips[node][ip] = cur_lastipcell
				cur_lastipcell = cur_lastipcell + 1
			if '|' in ips[ip][0]:
				if ips[ip][0].split('|')[1] == "Exclude" and IPrange == range:
					sheets[node]["C"+str(nodeips[node][ip])] = "Exclude"
				
	return nodeips 


def createSSHClient(loginIP, loginUser, loginPWD, ncpath):
	print("Create SSHClient login")
	sshclient = paramiko.SSHClient()
	sshclient.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	sshclient.connect(loginIP, username=loginUser, password=base64.b64decode(loginPWD))
	print("SSHClient connected to "+str(loginIP))
	stdin, stdout, stderr = sshclient.exec_command("if [ -f "+ncpath+"nc ]; then echo YES; else  echo NO ; fi")
	response = stdout.read().splitlines()[0].decode("ascii")
	print(response)
	if response == "NO":
		print("need to copy nc command to remote machine")
		# SCPCLient takes a paramiko transport as its only argument
		scpclient = scp.SCPClient(sshclient.get_transport())
		scpclient.put('./nc.zip', ncpath)
		scpclient.close()
		sshclient.exec_command("unzip "+ncpath+"nc.zip -d "+ncpath)
		sshclient.exec_command("chmod +x "+ncpath+"nc")
	return sshclient

def closeSSHClient(sshclient):
	if sshclient:
		sshclient.close()
		print("SSHClient disconnected")

def checkConnection(sshclient, node, nodeips, sheet,ipinfo, idx, ncpath):
	def updateLastCheck(sheet, idx):
		sheet["E"+str(idx)] = sheet["B"+str(idx)].value if sheet["B"+str(idx)].value != None else ''
		sheet["F"+str(idx)] = sheet["C"+str(idx)].value if sheet["C"+str(idx)].value != None else ''
		sheet["G"+str(idx)] = sheet["D"+str(idx)].value if sheet["D"+str(idx)].value != None else ''
	
	def getcurlExitStatus(std):
		print("GetCurlExitStatus")
		exitstatus = ""
		for line in std.read().splitlines():
			if "curl:" in line.decode("ascii"):
				exitstatus = line.decode("ascii")
				break
		
		if exitstatus == "":
			exitstatus = 'curl: (0) connect() OK'
		return exitstatus
		
	print("Check if connection is working")
	print(nodeips)
	for ip in nodeips:
		updateLastCheck(sheet, str(nodeips[ip]))
		print("Checking IP "+str(ip))
		if str(sheet["C"+str(nodeips[ip])].value) == 'Exclude' or str(sheet["F"+str(nodeips[ip])].value) == 'Exclude':
			statuscode = 'Exclude'
		elif str(sheet["F"+str(nodeips[ip])].value)[:8] != 'Success:':
			print("Check IP via curl command")
			#stdin, stdout, stderr = sshclient.exec_command("curl --connect-timeout 1 --max-time 1 "+ str(ip))
			#stdin, stdout, stderr = sshclient.exec_command("/home/ediadmin/nc -zv -w5 "+str(ip).split(':')[0]+" "+str(ip).split(':')[1]+" >/dev/null 2>&1 && echo $? || echo $? ")
			#check if nc command is on the remote machine
			
			stdin, stdout, stderr = sshclient.exec_command(ncpath+"nc -zv -w5 "+str(ip).split(':')[0]+" "+str(ip).split(':')[1]+" && echo $? || echo $? ")
			'''if stdout.read() == b'':
				statuscode = getcurlExitStatus(stderr)
				#statuscode = stderr.read().splitlines()[1].decode('ascii')
			else:
				statuscode = 'curl: (0) connect() OK'
			'''
			response = stderr.read().splitlines()
			response = response + stdout.read().splitlines()

			#response = stdout.read().splitlines()
			if response[1].decode("ascii") == '0':
				statuscode = 'Success: '+response[0].decode("ascii")
			else:
				statuscode = 'Failed: '+response[0].decode("ascii")
				
			#print(stdout.read())
			#print(stderr.read())
			#stdin, stdout, stderr = sshclient.exec_command("echo $?")
			#statuscode = stdout.read().splitlines()[0].decode('ascii')
			#print(statuscode)
		else:
			print("IP already tested successfully last time, no need to check again")
			statuscode = str(sheet["F"+str(nodeips[ip])].value)
			#ipinfo[ip][idx+5] = 'curl: (0) connect() OK'
			#sheet["C"+str(idx2+2)] = 'curl: (0) connect() OK'
			#sheet["D"+str(idx2+2)] = ''
			
		ipinfo[ip][idx+5] = statuscode
		sheet["B"+str(nodeips[ip])] = str(datetime.datetime.now().strftime("%y/%m/%d-%H:%M"))
		sheet["C"+str(nodeips[ip])] = statuscode
		sheet["D"+str(nodeips[ip])] = ''
		print("Statuscode was "+str(statuscode)+", also updated to ExcelSheet")
	
	return ipinfo


def updateIPSummary(sheet, ipinfo):
	print("UpdateSummary")
	print(ipinfo)
	loop = "true"
	idx = 2
	nodecount = len(next (iter (ipinfo.values())))-5
	print(nodecount)
	while loop == "true":
		cur_ip = str(sheet["B"+str(idx)].value)+':'+str(sheet["C"+str(idx)].value)
		for idx2 in range(nodecount):
			sheet[openpyxl.utils.get_column_letter(idx2+7)+str(idx)] = ipinfo[cur_ip][idx2+5]
		
		idx = idx + 1
		
		
		if sheet["A"+str(idx)].value == None:
			loop="false"
	
	print("UpdateSummary done")
	
def main():
	#Read Workbook and Worksheet
	excelsheetpath = getWorkBookFile()
	excelsheet = getWorkbook(excelsheetpath)
	sheets = getSheets(excelsheet)
	
	#Read Nodes and IPs
	nodeinfo = readNodes(sheets["Nodes"])
	ipinfo = readIPs(sheets["IPs"],len(nodeinfo))
	sheets = writeNodes(excelsheet, sheets, nodeinfo)
	nodesip = writeNodesIPs(sheets,ipinfo,nodeinfo)
	print(nodesip)
	#Execute check
	for idx,node in enumerate(nodeinfo):
		sshclient = createSSHClient(nodeinfo[node][2],nodeinfo[node][3],nodeinfo[node][4],nodeinfo[node][6])
		ipinfo = checkConnection(sshclient,node,nodesip[node],sheets[node],ipinfo, idx, nodeinfo[node][6])
		closeSSHClient(sshclient)
	
	updateIPSummary(sheets["IPs"], ipinfo)
	excelsheet.save(excelsheetpath)


#if script get called via command line/direct
if __name__ == "__main__":
    main()
